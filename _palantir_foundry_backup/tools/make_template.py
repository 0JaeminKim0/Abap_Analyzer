"""
작업 1 — 조직 정책별 Severity 수집 Excel Template 생성기

실행:
    pip install openpyxl
    python tools/make_template.py            # -> severity_policy_template.xlsx
    python tools/make_template.py out.xlsx    # 경로 지정

산출물(.xlsx)을 조직(아키텍트/보안팀)이 채운 뒤 Foundry에 업로드하면
`severity_policy_raw` Dataset 이 되고, transforms/policy_clean.py 가 정제한다.

시트 구성:
    Policy  : 19개 규칙 프리필. org_severity / enabled / owner / note 만 입력.
    README  : 작성 규칙.
"""
import os
import sys

# core/catalog.py 를 단일 원천으로 사용
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "src"))
from core.catalog import RULES  # noqa: E402

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

# 고정(잠금) 컬럼과 입력 컬럼 구분
FIXED_COLS = ["rule_id", "category", "rule_name", "default_severity"]
INPUT_COLS = ["org_severity", "enabled", "owner", "note"]
HEADERS = FIXED_COLS + INPUT_COLS

HEADER_FILL = PatternFill("solid", fgColor="0B7285")
FIXED_FILL = PatternFill("solid", fgColor="F1F3F5")
INPUT_FILL = PatternFill("solid", fgColor="FFF9DB")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="DEE2E6")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _build_policy_sheet(ws):
    ws.title = "Policy"
    ws.append(HEADERS)
    for ci, _ in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=ci)
        c.font = WHITE
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    for r in RULES:
        ws.append([
            r["rule_id"], r["category"], r["rule_name"], r["default_severity"],
            r["default_severity"],  # org_severity 기본값 = 제안값(조직이 조정)
            "Y", "", "",
        ])

    n = len(RULES)
    # 드롭다운: org_severity(E), enabled(F)
    dv_sev = DataValidation(type="list", formula1='"high,medium,low,off"', allow_blank=False)
    dv_sev.error = "high / medium / low / off 중 하나를 선택하세요."
    dv_sev.prompt = "조직 정책 등급. off = 규칙 비활성."
    ws.add_data_validation(dv_sev)
    dv_sev.add(f"E2:E{n + 1}")

    dv_en = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    dv_en.prompt = "이 규칙을 사용할지 여부."
    ws.add_data_validation(dv_en)
    dv_en.add(f"F2:F{n + 1}")

    # 스타일 + 셀 배경 (고정/입력 구분)
    for row in range(2, n + 2):
        for col in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=(HEADERS[col - 1] == "note"))
            cell.fill = FIXED_FILL if HEADERS[col - 1] in FIXED_COLS else INPUT_FILL

    widths = {"rule_id": 12, "category": 14, "rule_name": 34, "default_severity": 16,
              "org_severity": 16, "enabled": 10, "owner": 16, "note": 40}
    for ci, name in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = widths[name]
    ws.freeze_panes = "A2"


def _build_readme_sheet(ws):
    ws.title = "README"
    lines = [
        ("ABAP Analyzer — Severity 정책 수집 템플릿", True),
        ("", False),
        ("작성 방법", True),
        ("1) Policy 시트의 노란색 컬럼만 입력합니다. 회색 컬럼(rule_id/category/rule_name/default_severity)은 수정하지 마세요.", False),
        ("2) org_severity : 조직이 적용할 등급을 드롭다운에서 선택 (high / medium / low / off).", False),
        ("   - off  = 이 규칙을 분석에서 제외 (Agent 프롬프트에 포함되지 않음).", False),
        ("3) enabled : Y=사용, N=미사용. N 이면 off 와 동일하게 제외됩니다.", False),
        ("4) owner : 이 규칙 정책의 담당자. note : 등급을 조정한 사유.", False),
        ("", False),
        ("등급 의미", True),
        ("high   : 운영 장애 또는 보안 사고로 직결.", False),
        ("medium : 성능 저하·유지보수성 악화 (당장 장애는 아님).", False),
        ("low    : 스타일·표준 위반 (동작에 영향 없음).", False),
        ("", False),
        ("업로드", True),
        ("작성 완료된 이 파일을 Foundry에 업로드하면 severity_policy_raw 데이터셋이 됩니다.", False),
        ("이후 transforms/policy_clean.py 가 정제하여 Analyzer 에이전트가 자동으로 반영합니다.", False),
    ]
    for text, is_head in lines:
        ws.append([text])
        if is_head:
            ws.cell(row=ws.max_row, column=1).font = BOLD
    ws.column_dimensions["A"].width = 100


def main(out_path="severity_policy_template.xlsx"):
    wb = Workbook()
    _build_policy_sheet(wb.active)
    _build_readme_sheet(wb.create_sheet())
    wb.save(out_path)
    print(f"생성 완료: {out_path}  ({len(RULES)}개 규칙 프리필)")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "severity_policy_template.xlsx")
