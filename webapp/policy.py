"""
Railway 웹 서비스 정책(severity) 관리.

각 모듈/팀이 템플릿(xlsx 또는 csv)에 org_severity/enabled 를 작성해 업로드하면
활성 정책이 교체되고, 이후 분석에 즉시 반영된다.

정책 행 형식: {"rule_id","category","rule_name","severity"}  (severity: high|medium|low)

영속성:
  POLICY_STORE_PATH 환경변수가 있으면 그 경로(JSON)에 저장/로드한다.
  Railway 는 컨테이너가 휘발성이므로, 재시작 후에도 유지하려면 Railway Volume 경로를
  POLICY_STORE_PATH 로 지정할 것. 미지정 시 인스턴스 메모리에만 유지(재시작 시 기본값).
"""
import csv
import io
import json
import os

from core.catalog import RULES

_ALLOWED = {"high", "medium", "low"}
STORE_PATH = os.environ.get("POLICY_STORE_PATH", "").strip()

# 업로드 템플릿/파일에서 읽는 컬럼
TEMPLATE_HEADERS = ["rule_id", "category", "rule_name",
                    "default_severity", "org_severity", "enabled", "owner", "note"]


def catalog_policy():
    """카탈로그 기본 severity 로 구성한 정책."""
    return _apply({})


def _apply(overrides):
    """
    overrides: {rule_id: {"severity": str|None, "enabled": bool}} → 활성 정책 행 리스트.
    enabled=False 또는 severity=off 는 제외. 잘못된 severity 는 기본값으로 대체.
    """
    rows = []
    for r in RULES:
        ov = overrides.get(r["rule_id"], {})
        if ov.get("enabled") is False:
            continue
        sev = (ov.get("severity") or r["default_severity"]).lower()
        if sev == "off":
            continue
        if sev not in _ALLOWED:
            sev = r["default_severity"]
        rows.append({"rule_id": r["rule_id"], "category": r["category"],
                     "rule_name": r["rule_name"], "severity": sev})
    return rows


# ---------- 업로드 파싱 ----------
def parse_csv(data: bytes):
    text = data.decode("utf-8-sig", errors="replace")
    overrides = {}
    for row in csv.DictReader(io.StringIO(text)):
        rid = (row.get("rule_id") or "").strip()
        if not rid:
            continue
        enabled = (row.get("enabled") or "Y").strip().upper()
        sev = (row.get("org_severity") or row.get("severity") or "").strip().lower()
        overrides[rid] = {"severity": sev or None, "enabled": enabled != "N"}
    return overrides


def parse_xlsx(data: bytes):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb["Policy"] if "Policy" in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def cell(r, name):
        i = idx.get(name)
        return r[i] if (i is not None and i < len(r)) else None

    overrides = {}
    for r in rows[1:]:
        rid = cell(r, "rule_id")
        if not rid:
            continue
        enabled = str(cell(r, "enabled") or "Y").strip().upper()
        sev = str(cell(r, "org_severity") or "").strip().lower()
        overrides[str(rid).strip()] = {"severity": sev or None, "enabled": enabled != "N"}
    return overrides


def parse_upload(filename: str, data: bytes):
    name = (filename or "").lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return parse_xlsx(data)
    if name.endswith(".csv"):
        return parse_csv(data)
    # 확장자 불명 시 xlsx 시그니처(PK)로 추정
    if data[:2] == b"PK":
        return parse_xlsx(data)
    return parse_csv(data)


def template_csv():
    """모듈이 작성할 CSV 템플릿(기본값 프리필)."""
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADERS)
    for r in RULES:
        w.writerow([r["rule_id"], r["category"], r["rule_name"],
                    r["default_severity"], r["default_severity"], "Y", "", ""])
    return buf.getvalue()


def template_xlsx():
    """모듈이 작성할 Excel 템플릿(드롭다운 포함) 바이트."""
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Policy"
    ws.append(TEMPLATE_HEADERS)
    for r in RULES:
        ws.append([r["rule_id"], r["category"], r["rule_name"],
                   r["default_severity"], r["default_severity"], "Y", "", ""])

    n = len(RULES)
    # org_severity = E열, enabled = F열
    dv_sev = DataValidation(type="list", formula1='"high,medium,low,off"', allow_blank=False)
    dv_sev.prompt = "조직 정책 등급. off = 규칙 비활성."
    ws.add_data_validation(dv_sev)
    dv_sev.add(f"E2:E{n + 1}")

    dv_en = DataValidation(type="list", formula1='"Y,N"', allow_blank=False)
    dv_en.prompt = "이 규칙을 사용할지 여부."
    ws.add_data_validation(dv_en)
    dv_en.add(f"F2:F{n + 1}")

    widths = [12, 14, 34, 16, 16, 10, 16, 40]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(ord("A") + i)].width = w
    ws.freeze_panes = "A2"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------- 활성 정책 저장소 (인스턴스 단위) ----------
_active = None
_source = "default"


def _save(rows, source):
    if not STORE_PATH:
        return
    try:
        with open(STORE_PATH, "w", encoding="utf-8") as f:
            json.dump({"source": source, "rows": rows}, f, ensure_ascii=False)
    except OSError:
        pass


def _load():
    if not STORE_PATH or not os.path.exists(STORE_PATH):
        return None
    try:
        with open(STORE_PATH, encoding="utf-8") as f:
            d = json.load(f)
        return d.get("rows"), d.get("source", "stored")
    except (OSError, ValueError):
        return None


def get_active():
    global _active, _source
    if _active is None:
        loaded = _load()
        if loaded and loaded[0]:
            _active, _source = loaded
        else:
            _active, _source = catalog_policy(), "default"
    return _active


def get_source():
    get_active()
    return _source


def set_from_overrides(overrides, source):
    global _active, _source
    _active = _apply(overrides)
    _source = source
    _save(_active, source)
    return _active


def reset():
    return set_from_overrides({}, "default")
